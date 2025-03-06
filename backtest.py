import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from typing import Tuple, Dict, List, Optional
import os
import json
from datetime import datetime
import logging
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from strategy import PairGridStrategy

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger('Backtest')

class Backtest:
    """
    回测框架，用于测试策略在历史数据上的表现
    """
    
    def __init__(
        self,
        strategy: PairGridStrategy,
        data: pd.DataFrame,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        output_dir: str = '../backtest'
    ):
        """
        初始化回测
        
        Args:
            strategy: 策略实例
            data: 历史数据
            start_date: 回测开始日期，格式为 'YYYY-MM-DD'
            end_date: 回测结束日期，格式为 'YYYY-MM-DD'
            output_dir: 输出目录
        """
        self.strategy = strategy
        self.data = data
        
        # 确保数据索引是日期类型
        if not isinstance(data.index, pd.DatetimeIndex):
            raise ValueError("数据索引必须是日期类型")
            
        # 过滤日期范围
        if start_date:
            self.data = self.data[self.data.index >= start_date]
        if end_date:
            self.data = self.data[self.data.index <= end_date]
            
        self.output_dir = output_dir
        
        # 确保输出目录存在
        os.makedirs(output_dir, exist_ok=True)
        
        # 回测结果
        self.results = []
        
        logger.info(f"回测初始化完成，数据范围: {self.data.index[0]} 至 {self.data.index[-1]}")
    
    def run(self) -> Dict:
        """
        运行回测
        
        Returns:
            回测结果摘要
        """
        logger.info("开始回测...")
        
        # 确保数据足够长
        if len(self.data) <= self.strategy.atr_period:
            raise ValueError(f"数据长度不足，至少需要 {self.strategy.atr_period + 1} 天")
            
        # 按日期遍历数据
        for i in range(self.strategy.atr_period, len(self.data)):
            # 获取当前日期的数据窗口
            current_window = self.data.iloc[:i+1]
            
            # 更新策略
            result = self.strategy.update(current_window)
            
            # 记录结果
            self.results.append(result)
            
            # 如果触发熔断，跳过当天剩余操作
            if result.get('status') == 'circuit_breaker_triggered':
                logger.warning(f"日期 {current_window.index[-1]} 触发熔断，跳过当天交易")
                # 第二天重置熔断状态
                if i < len(self.data) - 1:
                    self.strategy.reset_circuit_breaker()
        
        # 获取策略表现摘要
        performance = self._calculate_performance_metrics()
        
        # 保存回测结果
        self._save_results()
        
        # 生成回测报告
        self._generate_report(performance)
        
        # 导出交易明细到Excel
        self._export_trades_to_excel()
        
        logger.info("回测完成")
        
        return performance
    
    def _calculate_performance_metrics(self) -> Dict:
        """
        计算详细的性能指标
        
        Returns:
            性能指标字典
        """
        # 提取每日组合价值
        portfolio_values = [result['portfolio_value'] for result in self.results if 'portfolio_value' in result]
        dates = [result['date'] for result in self.results if 'date' in result]
        
        if not portfolio_values:
            return {"status": "no_data"}
            
        initial_value = portfolio_values[0]
        final_value = portfolio_values[-1]
        
        # 计算累计收益率
        total_return = (final_value - initial_value) / initial_value
        
        # 计算日收益率
        daily_returns = []
        for i in range(1, len(portfolio_values)):
            daily_return = (portfolio_values[i] - portfolio_values[i-1]) / portfolio_values[i-1]
            daily_returns.append(daily_return)
            
        # 计算年化收益率 (假设252个交易日)
        trading_days = len(portfolio_values)
        years = trading_days / 252
        
        if years > 0:
            annualized_return = (1 + total_return) ** (1 / years) - 1
        else:
            annualized_return = 0
            
        # 计算波动率
        if daily_returns:
            # 过滤掉NaN和无穷大的值
            valid_returns = [r for r in daily_returns if not np.isnan(r) and not np.isinf(r)]
            if valid_returns:
                volatility = np.std(valid_returns)
                annualized_volatility = volatility * np.sqrt(252)
            else:
                volatility = 0
                annualized_volatility = 0
        else:
            volatility = 0
            annualized_volatility = 0
            
        # 计算夏普比率 (假设无风险利率为0)
        risk_free_rate = 0
        if annualized_volatility > 0:
            sharpe_ratio = (annualized_return - risk_free_rate) / annualized_volatility
        else:
            # 如果波动率为0，但收益率为正，则设置一个高夏普比率
            if annualized_return > risk_free_rate:
                sharpe_ratio = 10.0  # 设置一个较高的值表示低风险高收益
            # 如果波动率为0，但收益率为负，则设置一个低夏普比率
            elif annualized_return < risk_free_rate:
                sharpe_ratio = -10.0  # 设置一个较低的值表示低风险低收益
            else:
                sharpe_ratio = 0  # 收益率等于无风险利率
            
        # 计算最大回撤
        max_drawdown = 0
        peak = portfolio_values[0]
        
        for value in portfolio_values:
            if value > peak:
                peak = value
            drawdown = (peak - value) / peak
            max_drawdown = max(max_drawdown, drawdown)
            
        # 计算交易频率
        trade_count = len(self.strategy.trades)
        if trading_days > 0:
            trades_per_day = trade_count / trading_days
        else:
            trades_per_day = 0
            
        # 计算盈亏比
        if self.strategy.trades:
            profitable_trades = [trade for trade in self.strategy.trades if 
                                (trade['direction'] == 'BUY' and self.data.loc[dates[-1], f"{trade['symbol']}_close"] > trade['price']) or
                                (trade['direction'] == 'SELL' and self.data.loc[dates[-1], f"{trade['symbol']}_close"] < trade['price'])]
            profit_ratio = len(profitable_trades) / trade_count if trade_count > 0 else 0
        else:
            profit_ratio = 0
            
        return {
            "initial_value": initial_value,
            "final_value": final_value,
            "total_return": total_return,
            "annualized_return": annualized_return,
            "volatility": volatility,
            "annualized_volatility": annualized_volatility,
            "sharpe_ratio": sharpe_ratio,
            "max_drawdown": max_drawdown,
            "trade_count": trade_count,
            "trades_per_day": trades_per_day,
            "profit_ratio": profit_ratio,
            "trading_days": trading_days,
            "years": years
        }
    
    def _save_results(self) -> None:
        """保存回测结果"""
        # 提取每日组合价值
        dates = [result['date'] for result in self.results if 'date' in result]
        portfolio_values = [result['portfolio_value'] for result in self.results if 'portfolio_value' in result]
        
        # 创建结果DataFrame
        results_df = pd.DataFrame({
            'date': dates,
            'portfolio_value': portfolio_values
        })
        
        # 设置日期为索引
        results_df.set_index('date', inplace=True)
        
        # 保存为CSV
        results_path = os.path.join(self.output_dir, 'backtest_results.csv')
        results_df.to_csv(results_path)
        
        # 保存交易记录
        trades_path = os.path.join(self.output_dir, 'trades.json')
        with open(trades_path, 'w') as f:
            # 将交易记录转换为可序列化的格式
            serializable_trades = []
            for trade in self.strategy.trades:
                trade_copy = trade.copy()
                if isinstance(trade_copy['timestamp'], datetime):
                    trade_copy['timestamp'] = trade_copy['timestamp'].strftime('%Y-%m-%d %H:%M:%S')
                serializable_trades.append(trade_copy)
                
            json.dump(serializable_trades, f, indent=4)
        
        logger.info(f"回测结果已保存至 {self.output_dir}")
    
    def _export_trades_to_excel(self) -> None:
        """
        导出交易明细到Excel文件
        """
        if not self.strategy.trades:
            logger.warning("没有交易记录，无法导出Excel")
            return
            
        # 创建Excel文件
        excel_path = os.path.join(self.output_dir, 'trade_details.xlsx')
        
        # 将交易记录转换为DataFrame
        trades_data = []
        for trade in self.strategy.trades:
            trade_copy = trade.copy()
            if isinstance(trade_copy['timestamp'], datetime):
                trade_copy['timestamp'] = trade_copy['timestamp'].strftime('%Y-%m-%d %H:%M:%S')
            trades_data.append(trade_copy)
            
        trades_df = pd.DataFrame(trades_data)
        
        # 重命名列以符合要求
        column_mapping = {
            'timestamp': '时间戳',
            'symbol': '品种代码',
            'direction': '交易方向',
            'price': '成交价格',
            'quantity': '成交数量',
            'trigger_type': '触发类型',
            'grid_level': '关联网格层',
            'atr_value': '当前ATR值',
            'volatility_weight_ratio': '波动率权重比例',
            'circuit_breaker_active': '熔断状态标记',
            'pre_trade_cash': '交易前现金余额',
            'post_trade_cash': '交易后现金余额',
            'position_changes': '持仓变动明细',
            'reason': '交易原因',
            'boil_price': 'BOIL价格',
            'kold_price': 'KOLD价格'
        }
        
        trades_df = trades_df.rename(columns=column_mapping)
        
        # 保存为Excel
        trades_df.to_excel(excel_path, index=False, sheet_name='交易明细')
        
        # 使用openpyxl美化Excel
        wb = openpyxl.load_workbook(excel_path)
        ws = wb['交易明细']
        
        # 设置列宽
        for i, column in enumerate(trades_df.columns):
            column_width = max(len(str(column)), trades_df[column].astype(str).map(len).max())
            ws.column_dimensions[get_column_letter(i+1)].width = column_width + 4
        
        # 设置标题行样式
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        header_font = Font(bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 设置数据行样式
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 保存美化后的Excel
        wb.save(excel_path)
        
        logger.info(f"交易明细已导出至 {excel_path}")
    
    def _generate_report(self, performance: Dict) -> None:
        """
        生成回测报告
        
        Args:
            performance: 策略表现摘要
        """
        # 提取每日组合价值
        dates = [result['date'] for result in self.results if 'date' in result]
        portfolio_values = [result['portfolio_value'] for result in self.results if 'portfolio_value' in result]
        
        # 创建图表
        plt.figure(figsize=(12, 8))
        
        # 绘制组合价值曲线
        plt.subplot(2, 1, 1)
        plt.plot(dates, portfolio_values, label='Portfolio Value')
        plt.title('Backtest Results')
        plt.xlabel('Date')
        plt.ylabel('Portfolio Value')
        plt.grid(True)
        plt.legend()
        
        # 格式化x轴日期
        plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
        plt.gca().xaxis.set_major_locator(mdates.MonthLocator())
        plt.gcf().autofmt_xdate()
        
        # 绘制每日收益率
        daily_returns = []
        for i in range(1, len(portfolio_values)):
            daily_return = (portfolio_values[i] - portfolio_values[i-1]) / portfolio_values[i-1]
            daily_returns.append(daily_return)
            
        plt.subplot(2, 1, 2)
        plt.plot(dates[1:], daily_returns, label='Daily Returns')
        plt.axhline(y=0, color='r', linestyle='-')
        plt.xlabel('Date')
        plt.ylabel('Daily Return')
        plt.grid(True)
        plt.legend()
        
        # 格式化x轴日期
        plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
        plt.gca().xaxis.set_major_locator(mdates.MonthLocator())
        plt.gcf().autofmt_xdate()
        
        # 保存图表
        plt.tight_layout()
        plt.savefig(os.path.join(self.output_dir, 'backtest_chart.png'))
        
        # 创建性能指标报告
        report = f"""
        # 回测报告
        
        ## 基本信息
        - 回测期间: {dates[0]} 至 {dates[-1]}
        - 初始资金: {self.strategy.initial_capital}
        - 交易天数: {performance['trading_days']}
        
        ## 性能指标
        
        | 指标 | 值 | 计算公式 |
        |------|------|------|
        | 最终价值 | {performance['final_value']:.2f} | 最后一个交易日的组合价值 |
        | 累计收益率 | {performance['total_return']:.2%} | (最终价值 - 初始价值) / 初始价值 |
        | 年化收益率 | {performance['annualized_return']:.2%} | (1 + 累计收益率) ^ (1 / 年数) - 1 |
        | 年化波动率 | {performance['annualized_volatility']:.2%} | 日收益率标准差 * sqrt(252) |
        | 夏普比率 | {performance['sharpe_ratio']:.2f} | (年化收益率 - 无风险利率) / 年化波动率 |
        | 最大回撤 | {performance['max_drawdown']:.2%} | (峰值 - 谷值) / 峰值的最大值 |
        | 交易次数 | {performance['trade_count']} | 总交易次数 |
        | 日均交易频率 | {performance['trades_per_day']:.2f} | 总交易次数 / 交易天数 |
        | 盈利交易比例 | {performance['profit_ratio']:.2%} | 盈利交易数 / 总交易数 |
        
        ## 策略参数
        - 资产对: ({self.strategy.asset_1}, {self.strategy.asset_2})
        - ATR周期: {self.strategy.atr_period}
        - 网格间距: ATR的 {self.strategy.grid_distance_pct:.1%}
        - 最大单边仓位: 总资金的 {self.strategy.max_single_position_pct:.1%}
        - 对冲偏离阈值: {self.strategy.hedge_deviation_threshold:.1%}
        - 再平衡周期: {self.strategy.rebalance_period}
        - 熔断阈值: {self.strategy.circuit_breaker_threshold:.1%}
        """
        
        # 保存报告
        with open(os.path.join(self.output_dir, 'backtest_report.md'), 'w') as f:
            f.write(report)
        
        logger.info(f"回测报告已生成至 {self.output_dir}")


def prepare_data(data_path: str, asset_pair: Tuple[str, str]) -> pd.DataFrame:
    """
    准备回测数据
    
    Args:
        data_path: 数据文件路径
        asset_pair: 资产对
        
    Returns:
        处理后的数据
    """
    # 读取数据
    data = pd.read_csv(data_path)
    
    # 确保日期列存在
    date_columns = ['date', 'Date', 'datetime', 'Datetime']
    date_col = None
    for col in date_columns:
        if col in data.columns:
            date_col = col
            break
            
    if not date_col:
        # 尝试使用Timestamp列
        if 'Timestamp' in data.columns:
            # 将时间戳转换为日期
            data['date'] = pd.to_datetime(data['Timestamp'], unit='s')
            date_col = 'date'
        else:
            raise ValueError("数据必须包含日期列")
        
    # 转换日期列为日期类型
    data[date_col] = pd.to_datetime(data[date_col])
    
    # 设置日期为索引
    data.set_index(date_col, inplace=True)
    
    # 确保数据包含所需的列
    required_columns = []
    for asset in asset_pair:
        asset_cols = [f'{asset}_open', f'{asset}_high', f'{asset}_low', f'{asset}_close', f'{asset}_volume']
        alt_cols = [f'{asset}Open', f'{asset}High', f'{asset}Low', f'{asset}Close', f'{asset}Volume']
        
        # 检查列是否存在，如果不存在则尝试替代列名
        for i, col in enumerate(asset_cols):
            if col not in data.columns and alt_cols[i] in data.columns:
                # 重命名列
                data[col] = data[alt_cols[i]]
                
        required_columns.extend(asset_cols)
        
    missing_columns = [col for col in required_columns if col not in data.columns]
    if missing_columns:
        raise ValueError(f"数据缺少以下列: {missing_columns}")
        
    return data


if __name__ == "__main__":
    # 示例用法
    asset_pair = ('BOIL', 'KOLD')
    initial_capital = 100000
    
    # 创建策略实例
    strategy = PairGridStrategy(
        asset_pair=asset_pair,
        initial_capital=initial_capital,
        atr_period=14,
        grid_distance_pct=0.5,
        max_single_position_pct=0.2,
        hedge_deviation_threshold=0.05,
        rebalance_period='W-FRI',
        circuit_breaker_threshold=0.15
    )
    
    # 准备数据
    data = prepare_data('../data/sample_data.csv', asset_pair)
    
    # 创建回测实例
    backtest = Backtest(
        strategy=strategy,
        data=data,
        start_date='2022-01-01',
        end_date='2022-12-31',
        output_dir='../backtest'
    )
    
    # 运行回测
    performance = backtest.run()
    
    # 打印性能指标
    print("\n性能指标:")
    print(f"总收益率: {performance['total_return']:.2%}")
    print(f"年化收益率: {performance['annualized_return']:.2%}")
    print(f"年化波动率: {performance['annualized_volatility']:.2%}")
    print(f"夏普比率: {performance['sharpe_ratio']:.2f}")
    print(f"最大回撤: {performance['max_drawdown']:.2%}")
    print(f"交易次数: {performance['trade_count']}") 